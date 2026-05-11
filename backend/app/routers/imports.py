from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.audit import log_audit
from app.db import get_db
from app.deps import get_current_user, require_roles
from app.models import (
    Category,
    Location,
    Part,
    PartLocationStock,
    StockTransaction,
    StockTransactionType,
    TechnicianZoneAssignment,
    User,
)
from app.security import get_password_hash
from app.sku import generate_system_sku

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


router = APIRouter(prefix="/api/import", tags=["import"])

TEMP_TECHNICIAN_PASSWORD = "Westernpumps@26"
STORE_A_LOCATION = "Store A"
TECHNICIAN_ROLE = "technician"


class ImportSummary(BaseModel):
    created: int
    updated: int = 0
    skipped: int
    failed: int
    errors: list[str] = Field(default_factory=list)


class TechnicianImportSummary(BaseModel):
    created_users: int
    updated_users: int
    created_zones: int
    skipped: int
    failed: int
    errors: list[str] = Field(default_factory=list)


@dataclass
class TechnicianZoneRow:
    technician_name: str
    region_label: str
    station_name: str
    client_code: str | None
    zone_order: int


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _is_region_row(values: list[str]) -> bool:
    if not values:
        return False
    first = values[0].upper()
    return "REGION" in first or "AREA" in first


def _is_counts_row(values: list[str]) -> bool:
    return bool(values) and all("SITE" in value.upper() for value in values if value)


def _clean_person_name(name: str) -> str:
    cleaned = re.sub(r"\(.*?\)", "", name).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def _is_technician_header(value: str) -> bool:
    cleaned = _clean_person_name(value)
    if not cleaned or any(ch.isdigit() for ch in cleaned):
        return False
    upper = cleaned.upper()
    blocked = {"STATION", "CLIENT", "SPAKI", "TEMK", "VEK"}
    if upper in blocked or "REGION" in upper or "AREA" in upper or "SITE" in upper:
        return False
    parts = [piece for piece in cleaned.split(" ") if piece]
    return len(parts) >= 2 and all(re.fullmatch(r"[A-Za-z'-]+", piece) for piece in parts)


def _tech_email(full_name: str) -> str:
    parts = [re.sub(r"[^A-Za-z]", "", piece).lower() for piece in _clean_person_name(full_name).split()]
    parts = [piece for piece in parts if piece]
    if len(parts) < 2:
        raise ValueError(f"Cannot derive email from technician name '{full_name}'")
    return f"{parts[0]}{parts[1]}@gmail.com"


def _get_or_create_location(db: Session, name: str) -> Location:
    existing = db.scalar(select(Location).where(Location.name == name).limit(1))
    if existing:
        return existing
    location = Location(name=name)
    db.add(location)
    db.flush()
    return location


def _apply_store_a_quantity(db: Session, part: Part, quantity: int, current_user: User, note: str) -> None:
    location = _get_or_create_location(db, STORE_A_LOCATION)
    stock = db.scalar(
        select(PartLocationStock)
        .where(PartLocationStock.part_id == part.id, PartLocationStock.location_id == location.id)
        .limit(1)
    )
    previous_quantity = int(part.quantity_on_hand or 0)
    delta = quantity - previous_quantity
    part.quantity_on_hand = quantity
    part.location_id = location.id

    if stock is None:
        stock = PartLocationStock(part_id=part.id, location_id=location.id, quantity_on_hand=quantity)
        db.add(stock)
    else:
        stock.quantity_on_hand = quantity

    if delta != 0:
        db.add(
            StockTransaction(
                part_id=part.id,
                created_by_user_id=current_user.id if current_user else None,
                transaction_type=StockTransactionType.ADJUST,
                quantity_delta=delta,
                movement_type="IMPORT_ADJUST",
                notes=note,
            )
        )


def _parse_store_sheet(workbook) -> list[tuple[str, int | None]]:
    records: list[tuple[str, int | None]] = []
    seen: set[tuple[str, int | None]] = set()

    for ws in workbook.worksheets:
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            first = _normalize_text(row[0] if row else None)
            if not first:
                continue
            upper = first.upper()
            if idx <= 2 and ("ITEM" in upper or "ITEMS NOT AVAILABLE" in upper):
                continue
            if upper in {"ITEM", "ACTUAL STOCKS"} or "ITEMS NOT AVAILABLE" in upper:
                continue
            qty = _coerce_int(row[1] if len(row) > 1 else None)
            key = (_normalize_key(first), qty)
            if key in seen:
                continue
            seen.add(key)
            records.append((first, qty))
    return records


def _get_or_create_category(db: Session, name: str) -> Category:
    existing = db.scalar(select(Category).where(Category.name == name).limit(1))
    if existing:
        return existing
    category = Category(name=name, is_active=True)
    db.add(category)
    db.flush()
    return category


def _looks_like_pricing_sheet(workbook) -> bool:
    ws = workbook.active
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        cells = [_normalize_text(value).upper() for value in row]
        if "CATEGORY" in cells and "MATERIAL DESCRIPTION" in cells:
            return True
    return False


def _parse_pricing_sheet(workbook) -> list[tuple[str, str, str | None, float | None]]:
    ws = workbook.active
    rows: list[tuple[str, str, str | None, float | None]] = []
    for row in ws.iter_rows(values_only=True):
        category = _normalize_text(row[1] if len(row) > 1 else None)
        material = _normalize_text(row[2] if len(row) > 2 else None)
        unit = _normalize_text(row[3] if len(row) > 3 else None) or None
        rate = _coerce_float(row[4] if len(row) > 4 else None)

        if not material:
            continue
        upper_material = material.upper()
        if upper_material == "MATERIAL DESCRIPTION":
            continue
        if not category or category.upper() == "CATEGORY":
            continue
        rows.append((category, material, unit, rate))
    return rows


def _import_pricing_inventory(workbook, db: Session, dry_run: bool) -> ImportSummary:
    rows = _parse_pricing_sheet(workbook)
    existing_parts = db.scalars(select(Part)).all()
    by_name_category = {
        (_normalize_key(part.name), int(part.category_id) if part.category_id else 0): part
        for part in existing_parts
    }

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for category_name, material_name, unit, rate in rows:
        normalized_name = _normalize_key(material_name)
        normalized_category = _normalize_key(category_name)
        if not normalized_name or not normalized_category:
            skipped += 1
            continue

        if dry_run:
            created += 1
            continue

        try:
            category = _get_or_create_category(db, category_name)
            existing = by_name_category.get((normalized_name, int(category.id)))
            if existing is None:
                part = Part(
                    sku=generate_system_sku(db),
                    name=material_name.strip(),
                    category_id=category.id,
                    unit_of_measure=unit,
                    unit_price=rate,
                    quantity_on_hand=0,
                    min_quantity=0,
                    tracking_type="BATCH",
                    is_active=True,
                )
                db.add(part)
                db.flush()
                by_name_category[(normalized_name, int(category.id))] = part
                created += 1
            else:
                existing.name = material_name.strip()
                existing.category_id = category.id
                existing.unit_of_measure = unit
                existing.unit_price = rate
                existing.is_active = True
                updated += 1
        except Exception as exc:  # pragma: no cover
            db.rollback()
            errors.append(f"{material_name}: {exc}")
            continue

    if not dry_run:
        db.commit()
    return ImportSummary(
        created=created,
        updated=updated,
        skipped=skipped,
        failed=len(errors),
        errors=errors[:50],
    )


def _import_store_inventory(workbook, db: Session, current_user: User, dry_run: bool) -> ImportSummary:
    rows = _parse_store_sheet(workbook)
    existing_parts = db.scalars(select(Part)).all()
    by_name = {_normalize_key(part.name): part for part in existing_parts}

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for name, qty in rows:
        normalized_name = _normalize_key(name)
        if not normalized_name:
            skipped += 1
            continue
        is_sheet1_style = qty is not None
        desired_qty = max(qty or 0, 0)
        desired_min = desired_qty + 1 if is_sheet1_style else 1
        existing = by_name.get(normalized_name)

        if dry_run:
            if existing is None:
                created += 1
            else:
                updated += 1
            continue

        note = f"Imported from Store A.xlsx for {STORE_A_LOCATION}"
        try:
            if existing is None:
                part = Part(
                    sku=generate_system_sku(db),
                    name=name.strip(),
                    description=None,
                    image_url=None,
                    unit_price=None,
                    quantity_on_hand=0,
                    min_quantity=desired_min,
                    tracking_type="BATCH",
                    unit_of_measure=None,
                    location_id=None,
                    supplier_id=None,
                )
                db.add(part)
                db.flush()
                by_name[normalized_name] = part
                created += 1
                if is_sheet1_style:
                    _apply_store_a_quantity(db, part, desired_qty, current_user, note)
                else:
                    part.quantity_on_hand = 0
            else:
                part = existing
                part.name = name.strip()
                if is_sheet1_style:
                    _apply_store_a_quantity(db, part, desired_qty, current_user, note)
                updated += 1

            part.min_quantity = max(int(part.min_quantity or 0), desired_min)
            part.is_active = True
        except Exception as exc:  # pragma: no cover
            db.rollback()
            errors.append(f"{name}: {exc}")
            continue

    if not dry_run:
        db.commit()

    return ImportSummary(
        created=created,
        updated=updated,
        skipped=skipped,
        failed=len(errors),
        errors=errors[:50],
    )


def _parse_technician_zones(workbook) -> list[TechnicianZoneRow]:
    ws = workbook.active
    current_region = ""
    active_columns: dict[int, str] = {}
    zone_order: defaultdict[str, int] = defaultdict(int)
    rows: list[TechnicianZoneRow] = []

    for row in ws.iter_rows(values_only=True):
        text_values = [_normalize_text(value) for value in row]
        compact_values = [value for value in text_values if value]
        if not compact_values:
            continue
        if _is_region_row(compact_values):
            current_region = compact_values[0]
            active_columns = {}
            continue
        if _is_counts_row(compact_values):
            active_columns = {}
            continue

        header_candidates = {
            idx: value
            for idx, value in enumerate(text_values)
            if _is_technician_header(value)
            and (idx + 1 >= len(text_values) or not _normalize_text(text_values[idx + 1]))
        }
        if header_candidates:
            active_columns = {idx: _clean_person_name(value) for idx, value in header_candidates.items()}
            continue

        if any(value.upper() == "STATION" for value in compact_values):
            continue

        for col_idx, technician_name in active_columns.items():
            station_name = _normalize_text(text_values[col_idx] if col_idx < len(text_values) else "")
            client_code = _normalize_text(text_values[col_idx + 1] if col_idx + 1 < len(text_values) else "")
            if not station_name:
                continue
            zone_order[technician_name] += 1
            rows.append(
                TechnicianZoneRow(
                    technician_name=technician_name,
                    region_label=current_region or "Unassigned Region",
                    station_name=station_name,
                    client_code=client_code or None,
                    zone_order=zone_order[technician_name],
                )
            )

    return rows


def _import_technician_workbook(workbook, db: Session, current_user: User, dry_run: bool) -> TechnicianImportSummary:
    parsed_rows = _parse_technician_zones(workbook)
    grouped: dict[str, list[TechnicianZoneRow]] = defaultdict(list)
    for row in parsed_rows:
        grouped[row.technician_name].append(row)

    created_users = 0
    updated_users = 0
    created_zones = 0
    skipped = 0
    errors: list[str] = []

    if dry_run:
        seen_emails = {user.email.lower() for user in db.scalars(select(User)).all()}
        for technician_name, zones in grouped.items():
            try:
                email = _tech_email(technician_name)
            except ValueError as exc:
                errors.append(str(exc))
                continue
            if email in seen_emails:
                updated_users += 1
            else:
                created_users += 1
            created_zones += len(zones)
        return TechnicianImportSummary(
            created_users=created_users,
            updated_users=updated_users,
            created_zones=created_zones,
            skipped=skipped,
            failed=len(errors),
            errors=errors[:50],
        )

    password_hash = get_password_hash(TEMP_TECHNICIAN_PASSWORD)

    for technician_name, zones in grouped.items():
        try:
            email = _tech_email(technician_name)
        except ValueError as exc:
            errors.append(str(exc))
            continue

        user = db.scalar(select(User).where(User.email == email).limit(1))
        if user is None:
            user = User(
                tenant_id=current_user.tenant_id if current_user else 1,
                email=email,
                full_name=technician_name,
                role=TECHNICIAN_ROLE,
                password_hash=password_hash,
                is_active=True,
                must_change_password=True,
            )
            db.add(user)
            db.flush()
            created_users += 1
        else:
            user.full_name = technician_name
            user.role = TECHNICIAN_ROLE
            user.is_active = True
            user.must_change_password = True
            updated_users += 1

        db.query(TechnicianZoneAssignment).filter(TechnicianZoneAssignment.user_id == user.id).delete()
        for zone in zones:
            db.add(
                TechnicianZoneAssignment(
                    tenant_id=user.tenant_id,
                    user_id=user.id,
                    region_label=zone.region_label,
                    station_name=zone.station_name,
                    client_code=zone.client_code,
                    zone_order=zone.zone_order,
                )
            )
            created_zones += 1

    db.commit()
    return TechnicianImportSummary(
        created_users=created_users,
        updated_users=updated_users,
        created_zones=created_zones,
        skipped=skipped,
        failed=len(errors),
        errors=errors[:50],
    )


def _load_workbook_from_upload(file: UploadFile):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    payload = file.file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded workbook is empty")
    try:
        return openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=400, detail=f"Failed to read workbook: {exc}") from exc


@router.post(
    "/inventory-xlsx",
    response_model=ImportSummary,
    dependencies=[Depends(require_roles("store_manager", "manager", "admin"))],
)
def import_inventory_xlsx(
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportSummary:
    workbook = _load_workbook_from_upload(file)
    if _looks_like_pricing_sheet(workbook):
        summary = _import_pricing_inventory(workbook, db, dry_run)
    else:
        summary = _import_store_inventory(workbook, db, current_user, dry_run)
    log_audit(
        db,
        current_user,
        action="import",
        entity_type="inventory",
        detail={
            "file_name": file.filename,
            "dry_run": dry_run,
            "created": summary.created,
            "updated": summary.updated,
            "skipped": summary.skipped,
            "failed": summary.failed,
        },
    )
    db.commit()
    return summary


@router.post(
    "/technicians-zones-xlsx",
    response_model=TechnicianImportSummary,
    dependencies=[Depends(require_roles("admin", "manager"))],
)
def import_technicians_zones_xlsx(
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> TechnicianImportSummary:
    workbook = _load_workbook_from_upload(file)
    summary = _import_technician_workbook(workbook, db, current_user, dry_run)
    log_audit(
        db,
        current_user,
        action="import",
        entity_type="technician_zones",
        detail={
            "file_name": file.filename,
            "dry_run": dry_run,
            "created_users": summary.created_users,
            "updated_users": summary.updated_users,
            "created_zones": summary.created_zones,
            "failed": summary.failed,
        },
    )
    db.commit()
    return summary
